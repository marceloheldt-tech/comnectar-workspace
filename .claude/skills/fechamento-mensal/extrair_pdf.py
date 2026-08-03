"""
Extrator de relatórios em PDF do Bling — comnéctar
Uso:
  python extrair_pdf.py --tipo vendas       --arquivo "financeiro/2026-07/vendas julho.pdf"       --saida "financeiro/2026-07/vendas.xlsx"
  python extrair_pdf.py --tipo pagamentos   --arquivo "financeiro/2026-07/pagamentos julho.pdf"   --saida "financeiro/2026-07/contas-pagas.xlsx"
  python extrair_pdf.py --tipo recebimentos --arquivo "financeiro/2026-07/recebimentos julho.pdf" --saida "financeiro/2026-07/contas-recebidas.xlsx"

O Bling exporta esses 3 relatórios como PDF, não Excel. Esse script lê a posição de
cada palavra na página (via pdfplumber) e reconstrói as linhas da tabela — mais
confiável que extrair o texto puro, que embaralha colunas quando o nome do produto
quebra em várias linhas.

- "vendas": relatório "por Produto" — cada linha tem Código, CFOP e uma sequência de
  valores terminando em Quantidade ... Valor Total, Percentual. Casa pelo código
  (não pelo nome, que fica cortado) e extrai Quantidade + Valor Total.
- "pagamentos"/"recebimentos": relatórios "por Categoria"/lista simples — cada linha é
  um rótulo (categoria ou cliente) seguido do valor. A linha "Total" é usada só pra
  conferir a soma, não entra nos dados.

Sempre imprime o total somado dos dados extraídos pra conferir contra o total impresso
no PDF — se não bater, alguma linha não foi capturada corretamente.
"""
import argparse
import re
import openpyxl
import pdfplumber

RE_CODIGO = re.compile(r"\d+")
RE_CFOP = re.compile(r"\d{4}")
RE_MONEY = re.compile(r"[\d\.]+,\d{2}")


def pv(s):
    return float(s.replace(".", "").replace(",", "."))


def linhas_por_posicao(pdf_path):
    todas = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            linhas = {}
            for w in words:
                top = round(w["top"] / 3) * 3
                linhas.setdefault(top, []).append(w)
            for top in sorted(linhas.keys()):
                todas.append(sorted(linhas[top], key=lambda w: w["x0"]))
    return todas


def extrair_vendas(pdf_path):
    registros = []
    for ws in linhas_por_posicao(pdf_path):
        if not ws:
            continue
        primeiro = ws[0]
        if not (25 <= primeiro["x0"] <= 40 and RE_CODIGO.fullmatch(primeiro["text"])):
            continue
        cfop_tok = next((w for w in ws if 235 <= w["x0"] <= 255 and RE_CFOP.fullmatch(w["text"])), None)
        if cfop_tok is None:
            continue
        depois_cfop = [w for w in ws if w["x0"] > cfop_tok["x0"]]
        numeros = [w["text"] for w in depois_cfop if RE_MONEY.fullmatch(w["text"])]
        if len(numeros) < 2:
            continue
        registros.append({
            "codigo": primeiro["text"],
            "quantidade": pv(numeros[0]),
            "valor_total": pv(numeros[-2]),
        })
    return registros


def extrair_lista_valor(pdf_path):
    registros = []
    total = None
    for ws in linhas_por_posicao(pdf_path):
        if not ws:
            continue
        ultimo = ws[-1]
        if not RE_MONEY.fullmatch(ultimo["text"]):
            continue
        rotulo = " ".join(w["text"] for w in ws[:-1]).strip()
        if not rotulo or rotulo.lower() == "valor":
            continue
        valor = pv(ultimo["text"])
        if rotulo.lower() == "total":
            total = valor
        else:
            registros.append({"rotulo": rotulo, "valor": valor})
    return registros, total


def salvar_vendas(registros, saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"
    ws.append(["Código", "Quantidade", "Valor Total"])
    for r in registros:
        ws.append([r["codigo"], r["quantidade"], r["valor_total"]])
    wb.save(saida)


def salvar_lista(registros, saida, coluna_rotulo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append([coluna_rotulo, "Valor"])
    for r in registros:
        ws.append([r["rotulo"], r["valor"]])
    wb.save(saida)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tipo", required=True, choices=["vendas", "pagamentos", "recebimentos"])
    ap.add_argument("--arquivo", required=True)
    ap.add_argument("--saida", required=True)
    args = ap.parse_args()

    if args.tipo == "vendas":
        registros = extrair_vendas(args.arquivo)
        soma_qtd = sum(r["quantidade"] for r in registros)
        soma_valor = sum(r["valor_total"] for r in registros)
        salvar_vendas(registros, args.saida)
        print(f"{len(registros)} produtos extraídos. Quantidade total: {soma_qtd:.2f} | Valor total: R$ {soma_valor:.2f}")
        print("Confira esses dois números contra a linha 'Totais' do PDF antes de seguir.")
    else:
        coluna = "Categoria" if args.tipo == "pagamentos" else "Descrição"
        registros, total_pdf = extrair_lista_valor(args.arquivo)
        soma = sum(r["valor"] for r in registros)
        salvar_lista(registros, args.saida, coluna)
        print(f"{len(registros)} linhas extraídas. Soma: R$ {soma:.2f}")
        if total_pdf is not None:
            bate = "OK" if abs(soma - total_pdf) < 0.01 else "DIVERGE"
            print(f"Total no PDF: R$ {total_pdf:.2f} — {bate}")
        else:
            print("Não encontrei a linha 'Total' no PDF pra conferir — revisar manualmente.")


if __name__ == "__main__":
    main()
