# -*- coding: utf-8 -*-
"""
Módulo reutilizável pra gerar o Excel de fluxo de conteúdo da comnéctar.
Não roda sozinho — importar `build_workbook` a partir de um script que define
a pauta (guia de imagens + conteúdo de cada semana) e chama a função.

Uso típico:

    import sys
    sys.path.insert(0, r"c:\Users\marce\Desktop\claude comnéctar\.claude\skills\planejamento-conteudo")
    from gerar_excel import build_workbook, GUIA_IMAGENS_PADRAO

    semanas = {
        "Semana 1": [
            ["Seg 7/set", "Instagram", "Diário da comnéctar", "Tema da peça", "Texto completo...", "Tipo de imagem sugerida"],
            ...
        ],
        "Semana 2": [...],
    }

    build_workbook(
        out_path=r"c:\Users\marce\Desktop\claude comnéctar\campanhas\planejamento-t1-2026-2027\conteudos-t1.xlsx",
        semanas=semanas,
    )

Cada linha de semana segue sempre a mesma ordem de colunas:
["Data", "Canal", "Formato/Série", "Tema", "Texto", "Imagem sugerida"]
Canal precisa ser exatamente "Instagram", "WhatsApp" ou "Email" pra colorir certo.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CANAL_COLORS = {
    "Instagram": "F3D6E0",
    "WhatsApp": "D9EAD3",
    "Email": "D0E3F0",
}

HEADER_FILL = "5B1F3A"
HEADER_FONT_COLOR = "FFFFFF"

CONTENT_HEADERS = ["Data", "Canal", "Formato/Série", "Tema", "Texto", "Imagem sugerida"]
CONTENT_COL_WIDTHS = [12, 12, 20, 26, 70, 34]

GUIA_IMAGENS_PADRAO = [
    [
        "Foto de produto (garrafa)",
        "Sugestão de vinho no WhatsApp, posts de oferta, destaque de rótulo",
        "Já vem pronta do Shopify ou dos PDFs dos fornecedores (mesmo processo usado nos catálogos)",
    ],
    [
        "Foto de bastidores / lifestyle",
        "Posts \"Diário da comnéctar\" e \"Bastidores da importação\"",
        "Precisa ser fotografada por você: adega, mesa de degustação, caixas chegando, você escolhendo vinhos. Não precisa ser produção, celular com boa luz resolve",
    ],
    [
        "Foto do produtor / vinícola",
        "\"Histórias dos produtores\", história de produtor no WhatsApp",
        "Pegar nos PDFs dos fornecedores (a Tanyno já tem material assim) ou no site/Instagram do próprio produtor, sempre dando crédito",
    ],
    [
        "Arte gráfica educativa (texto sobre fundo de marca, sem foto)",
        "Posts educativos como \"como ler um rótulo\" e \"vale ou não vale\"",
        "Não precisa de foto real. Dá pra gerar com a skill /carrossel ou /post-produto usando o design guide da marca",
    ],
    [
        "Print ou foto enviada pelo cliente",
        "Prova social, depoimentos, semanas de fechamento de trimestre",
        "Pedir print de conversa (com autorização) ou foto que o próprio cliente mandou brindando com o vinho",
    ],
]


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.font = Font(color=HEADER_FONT_COLOR, bold=True, size=11)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _add_content_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(sheet_name)
    ws.append(CONTENT_HEADERS)
    _style_header(ws, len(CONTENT_HEADERS))

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in rows:
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        canal = ws.cell(row=r, column=2).value
        fill_color = CANAL_COLORS.get(canal, "FFFFFF")
        for c in range(1, len(CONTENT_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            cell.border = border
            if c == 2:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(bold=True)
        ws.row_dimensions[r].height = 90

    for i, w in enumerate(CONTENT_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def _add_guia_imagens_sheet(wb, guia_rows):
    ws = wb.create_sheet("Guia de imagens")
    headers = ["Tipo de imagem", "Onde é usada", "Como conseguir"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for row in guia_rows:
        ws.append(row)

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            cell.border = border
        ws.row_dimensions[r].height = 60

    widths = [30, 40, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def build_workbook(out_path, semanas, guia_rows=None):
    """
    semanas: dict ordenado {"Semana 1": [[...linha...], ...], "Semana 2": [...], ...}
    guia_rows: lista de linhas pra aba "Guia de imagens" (usa GUIA_IMAGENS_PADRAO se None)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_guia_imagens_sheet(wb, guia_rows if guia_rows is not None else GUIA_IMAGENS_PADRAO)

    for nome_semana, rows in semanas.items():
        _add_content_sheet(wb, nome_semana, rows)

    wb.save(out_path)
    return out_path
